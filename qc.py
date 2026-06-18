"""
qc.py — QC checklist cross-check engine.

Given:
  - the reference data already extracted from the main agreement PDF
    (the dict produced by extract_with_groq in main.py), and
  - a set of uploaded supporting files (PDFs and images), and
  - the rules in rules.yaml,
this module verifies each checklist row and fills the EXACT uploaded
QC template, returning the populated .xlsx bytes plus a result list
for on-page display.
"""
import os
import io
import re
import json
import base64

import yaml
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Font
from groq import Groq

GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
GROQ_TEXT_MODEL = os.environ.get("GROQ_MODEL", "llama-3.3-70b-versatile")
GROQ_VISION_MODEL = os.environ.get("GROQ_VISION_MODEL", "meta-llama/llama-4-scout-17b-16e-instruct")

# Which template sheet to fill, and the columns to write into.
TEMPLATE_SHEET = os.environ.get("QC_TEMPLATE_SHEET", "Sheet1")
COL_YESNO = 3   # column C
COL_REMARK = 4  # column D (merged D:E)


# ---------------------------------------------------------------------------
# Rules
# ---------------------------------------------------------------------------
def load_rules(path: str = "rules.yaml") -> list:
    with open(path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    return data.get("rules", [])


# ---------------------------------------------------------------------------
# Reference field mapping
# ---------------------------------------------------------------------------
def _build_reference(ref_data: dict) -> dict:
    """
    Map the agreement-extraction dict into the simple reference fields the
    rules compare against (panel_model, panel_qty, inverter_model, etc.).
    Pulls model/qty out of the line_items where possible.
    """
    ref = {
        "customer_name": ref_data.get("customer_name", ""),
        "billing_address": ref_data.get("billing_address", ""),
        "panel_model": "",
        "panel_qty": "",
        "inverter_model": "",
        "battery_model": "",
    }
    for li in ref_data.get("line_items") or []:
        item = (li.get("item") or "").lower()
        spec = li.get("specification") or ""
        qty = li.get("quantity") or ""
        if "panel" in item:
            ref["panel_model"] = ref["panel_model"] or spec
            ref["panel_qty"] = ref["panel_qty"] or qty
        elif "inverter" in item and "phase" not in item:
            ref["inverter_model"] = ref["inverter_model"] or spec
        elif "battery" in item:
            ref["battery_model"] = ref["battery_model"] or spec
    return ref


# ---------------------------------------------------------------------------
# Source reading
# ---------------------------------------------------------------------------
def _pdf_text(file_bytes: bytes) -> str:
    parts = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


def _client() -> Groq:
    if not GROQ_API_KEY:
        raise RuntimeError("GROQ_API_KEY not set. Add it to your .env file.")
    return Groq(api_key=GROQ_API_KEY)


def _extract_from_pdf(file_bytes: bytes, want: str) -> str:
    text = _pdf_text(file_bytes)
    prompt = (
        f"From the document text below, extract {want}. "
        "Reply with ONLY the extracted value, nothing else. "
        "If it is not present, reply exactly with NOT_FOUND.\n\n"
        f'"""\n{text[:12000]}\n"""'
    )
    resp = _client().chat.completions.create(
        model=GROQ_TEXT_MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=0,
    )
    return (resp.choices[0].message.content or "").strip()


def _extract_from_image(file_bytes: bytes, mime: str, want: str) -> str:
    b64 = base64.b64encode(file_bytes).decode("ascii")
    prompt = (
        f"Look at this image and extract {want}. "
        "Reply with ONLY the extracted value, nothing else. "
        "If it is not present, reply exactly with NOT_FOUND."
    )
    resp = _client().chat.completions.create(
        model=GROQ_VISION_MODEL,
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url",
                 "image_url": {"url": f"data:{mime};base64,{b64}"}},
            ],
        }],
        temperature=0,
    )
    return (resp.choices[0].message.content or "").strip()


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------
def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _digits(s: str) -> str:
    m = re.search(r"\d+(?:\.\d+)?", s or "")
    return m.group(0).rstrip("0").rstrip(".") if m else ""


def _compare(source_val: str, ref_val: str, match: str):
    """Return (passed: bool, detail: str)."""
    if not source_val or source_val.upper() == "NOT_FOUND":
        return False, "value not found in source"
    if not ref_val:
        return False, "no reference value to compare"

    if match == "numeric":
        a, b = _digits(source_val), _digits(ref_val)
        return (a == b and a != ""), f"source={source_val} | reference={ref_val}"
    if match == "contains":
        ns, nr = _norm(source_val), _norm(ref_val)
        substr = nr in ns or ns in nr
        # token overlap: every alphanumeric token of the shorter value appears
        # in the longer one (handles "FOXESS KH10" vs "FOXESS inverter KH10").
        ts = set(re.findall(r"[a-z0-9]+", source_val.lower()))
        tr = set(re.findall(r"[a-z0-9]+", ref_val.lower()))
        shorter, longer = (ts, tr) if len(ts) <= len(tr) else (tr, ts)
        token_ok = bool(shorter) and shorter.issubset(longer)
        return (substr or token_ok), f"source='{source_val}' | reference='{ref_val}'"
    # exact
    ok = _norm(source_val) == _norm(ref_val)
    return ok, f"source='{source_val}' | reference='{ref_val}'"


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def run_crosscheck(ref_data: dict, sources: dict, rules_path: str = "rules.yaml"):
    """
    ref_data : dict from the agreement extraction (main.py extract_with_groq)
    sources  : {filename: {"bytes": b"...", "mime": "application/pdf"|"image/..."}}
    Returns  : (results: list[dict], reference: dict)
      each result: {row, label, type, status ('Y'/'N'), remark}
    """
    rules = load_rules(rules_path)
    reference = _build_reference(ref_data)
    results = []

    for rule in rules:
        row = rule.get("row")
        label = rule.get("label", "")
        rtype = rule.get("type", "compare")
        src_name = rule.get("source", "")
        src = sources.get(src_name)

        # Source missing entirely.
        if src is None:
            results.append({
                "row": row, "label": label, "type": rtype,
                "status": "N", "remark": f"Source '{src_name}' not uploaded",
            })
            continue

        if rtype == "presence":
            results.append({
                "row": row, "label": label, "type": rtype,
                "status": "Y", "remark": f"Provided: {src_name}",
            })
            continue

        # compare
        want = rule.get("source_field", "the relevant value")
        ref_field = rule.get("reference_field", "")
        ref_val = reference.get(ref_field, "")
        try:
            if rule.get("is_image"):
                src_val = _extract_from_image(src["bytes"], src["mime"], want)
            else:
                src_val = _extract_from_pdf(src["bytes"], want)
        except Exception as e:
            results.append({
                "row": row, "label": label, "type": rtype,
                "status": "N", "remark": f"Read error: {e}",
            })
            continue

        passed, detail = _compare(src_val, ref_val, rule.get("match", "exact"))
        results.append({
            "row": row, "label": label, "type": rtype,
            "status": "Y" if passed else "N",
            "remark": ("Match. " if passed else "MISMATCH. ") + detail,
        })

    return results, reference


def fill_template(template_bytes: bytes, results: list, ref_data: dict) -> bytes:
    """
    Load the EXACT uploaded QC template and write only the Yes/No (col C) and
    Remarks (col D) cells for each result row, plus the header. Layout, merged
    cells, fonts and widths are all preserved by openpyxl.
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb[TEMPLATE_SHEET] if TEMPLATE_SHEET in wb.sheetnames else wb.active

    # Fill header fields (append to the existing label text in column A).
    def set_header(row, prefix, value):
        if value:
            cell = ws.cell(row=row, column=1)
            base = (cell.value or prefix)
            cell.value = f"{base.rstrip()} {value}" if not str(base).strip().endswith(str(value)) else base

    set_header(3, "Customer Name  :", ref_data.get("customer_name", ""))
    set_header(4, "Correct Address  :", ref_data.get("billing_address", ""))

    for r in results:
        row = r["row"]
        if not row:
            continue
        ws.cell(row=row, column=COL_YESNO).value = r["status"]
        ws.cell(row=row, column=COL_REMARK).value = r["remark"]

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
