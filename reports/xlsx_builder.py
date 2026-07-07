"""
reports/xlsx_builder.py — parse a QC checklist .xlsx into structured items,
and regenerate a styled .xlsx from items (used for download after editing).
"""
import io
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

SECTION_HINT = re.compile(r"DETAILS|CEC|APPROVED|SYSTEM|CUSTOMER", re.IGNORECASE)

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_xlsx(file_bytes: bytes, sheet_name: str = None):
    """
    Read a QC checklist sheet and return (items, header_labels, note_text).
    items: list of {position, sno, is_section, text, active}
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        def score(s):
            return sum(1 for r in range(1, s.max_row + 1) if s.cell(r, 4).value)
        ws = max(wb.worksheets, key=score)

    note_text = ""
    items = []
    pos = 0

    start_row = None
    col_map = {"sno": 1, "text": 2, "yn": 3, "ref": 4, "prompt": 5}

    def _is_sno(v):
        return (v.startswith("sno") or v.startswith("s.no") or
                v.startswith("sr.no") or v in ("s no", "sr no", "#", "no."))

    def _is_checklist(v):
        return "checklist" in v

    for r in range(1, ws.max_row + 1):
        row_vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, 15)]
        has_sno = any(_is_sno(v) for v in row_vals)
        has_checklist = any(_is_checklist(v) for v in row_vals)
        if has_sno and has_checklist:
            start_row = r + 1
            for c, v in enumerate(row_vals, start=1):
                if _is_sno(v):
                    col_map["sno"] = c
                elif _is_checklist(v):
                    col_map["text"] = c
                elif "yes" in v and "no" in v:
                    col_map["yn"] = c
                elif "what to verify" in v or "verify as per" in v or "prompt" in v:
                    col_map["prompt"] = c
                elif "remark" in v or "reference" in v:
                    col_map["ref"] = c
            break

    if start_row is None:
        start_row = 9

    note_cell = ws.cell(1, col_map["ref"]).value
    if note_cell and "note" not in str(note_cell).lower():
        note_text = str(note_cell)

    for r in range(start_row, ws.max_row + 1):
        sno = ws.cell(r, col_map["sno"]).value
        text = ws.cell(r, col_map["text"]).value
        if text is None and sno is None:
            continue
        text = str(text).strip() if text is not None else ""
        sno_str = str(sno).strip() if sno is not None else ""
        if not text:
            continue

        remark = ws.cell(r, col_map["ref"]).value
        reference = ""
        if remark is not None:
            reference = re.sub(r"^\s*Refer to\s+", "", str(remark).strip(),
                               flags=re.IGNORECASE).strip()

        prompt_val = ws.cell(r, col_map["prompt"]).value
        prompt_str = str(prompt_val).strip() if prompt_val is not None else ""

        pos += 1
        is_section = (
            bool(text.isupper() and SECTION_HINT.search(text)) or
            (text.isupper() and len(text) > 4 and not sno_str.isdigit())
        )

        items.append({
            "position": pos,
            "sno": sno_str,
            "is_section": is_section,
            "text": text,
            "reference": reference,
            "prompt": prompt_str,
            "active": 1,
        })

    header_labels = {
        "customer_label": _label_at(ws, "customer"),
        "address_label": _label_at(ws, "address"),
        "job_label": _label_at(ws, "job details"),
    }
    return items, header_labels, note_text


def _label_at(ws, needle: str):
    for r in range(1, 12):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "")
            if needle.lower() in v.lower():
                return v.split(":")[0].strip() + "  :"
    return ""


def build_xlsx(items: list, header_labels: dict = None, note_text: str = "",
               title: str = "||  JOB COMPLIANCE CHECKLIST (QC)  ||",
               filled: dict = None, email_results: list = None) -> bytes:
    """
    Build a styled .xlsx from structured items.
    filled (optional): {position: {"status": "Y"/"N", "remark": "..."}}
    """
    header_labels = header_labels or {}
    filled = filled or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    bold = Font(name="Calibri", size=11, bold=True)
    normal = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    section_fill  = PatternFill("solid", fgColor="E8590C")
    section_font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    changed_fill  = PatternFill("solid", fgColor="FFE58A")

    ws.merge_cells("A1:E1")
    c = ws.cell(1, 1, title)
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A3:C3")
    ws.cell(3, 1, header_labels.get("customer_label") or "Customer Name  :").font = bold
    ws.merge_cells("A4:C4")
    ws.cell(4, 1, header_labels.get("address_label") or "Correct Address  :").font = bold
    ws.merge_cells("A5:C5")
    ws.cell(5, 1, "Checked By  :").font = bold
    ws.merge_cells("A6:C6")
    ws.cell(6, 1, "Date  :").font = bold
    ws.merge_cells("A7:E7")
    ws.cell(7, 1, header_labels.get("job_label") or "Job Details  :").font = bold

    headers = ["Sno.", "Checklist Item", "Yes/No", "Remarks", "What to verify as per Agreement"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(8, ci, h)
        cell.font = bold
        cell.alignment = center
        cell.border = BORDER
    ws.row_dimensions[8].height = 20

    r = 9
    MIN_ROW_HEIGHT = 30

    # ── Email Verification section (only when Step 2 results are present) ──────
    if email_results:
        # Section header row
        sec_cell = ws.cell(r, 2, "EMAIL VERIFICATION")
        sec_cell.font  = section_font
        sec_cell.fill  = section_fill
        for ci in (1, 2, 3, 4, 5):
            c = ws.cell(r, ci)
            c.fill   = section_fill
            c.border = BORDER
        ws.cell(r, 1).font = section_font
        ws.cell(r, 3).font = section_font
        ws.cell(r, 4).font = section_font
        ws.cell(r, 5).font = section_font
        r += 1

        for idx, er in enumerate(email_results, start=1):
            name      = er.get("name", "")
            status    = er.get("status", "N/A")
            ai_status = er.get("ai_status", "")
            remark    = er.get("remark", "")

            sno_c    = ws.cell(r, 1, str(idx))
            text_c   = ws.cell(r, 2, name)
            yn_c     = ws.cell(r, 3, status)
            remark_c = ws.cell(r, 4, remark)
            prompt_c = ws.cell(r, 5, "Email attachment verification")

            for cell in (sno_c, text_c, yn_c, remark_c, prompt_c):
                cell.border = BORDER

            sno_c.font    = bold
            sno_c.alignment  = center
            text_c.font   = normal
            text_c.alignment = left
            yn_c.alignment   = center
            remark_c.alignment = left
            prompt_c.alignment = left

            if ai_status and status != ai_status:
                yn_c.fill = changed_fill

            ws.row_dimensions[r].height = MIN_ROW_HEIGHT
            r += 1

    for it in items:
        if not it.get("active", 1):
            continue
        pos = it.get("position")
        is_section = it.get("is_section")

        sno_cell    = ws.cell(r, 1, it.get("sno", ""))
        text_cell   = ws.cell(r, 2, it.get("text", ""))
        yn_cell     = ws.cell(r, 3, "")
        remark_cell = ws.cell(r, 4, "")
        prompt_cell = ws.cell(r, 5, it.get("prompt", ""))

        for cell in (sno_cell, text_cell, yn_cell, remark_cell, prompt_cell):
            cell.border = BORDER

        if is_section:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
            text_cell.font   = section_font
            text_cell.fill   = section_fill
            sno_cell.fill    = section_fill
            yn_cell.fill     = section_fill
            remark_cell.fill = section_fill
            prompt_cell.fill = section_fill
        else:
            sno_cell.font  = bold
            text_cell.font = normal

        sno_cell.alignment    = center
        text_cell.alignment   = left
        yn_cell.alignment     = center
        remark_cell.alignment = left
        prompt_cell.alignment = left

        fv = filled.get(pos)
        if fv:
            yn_cell.value     = fv.get("status", "")
            remark_cell.value = fv.get("remark", "")
            ai_status = fv.get("ai_status", "")
            if ai_status and fv.get("status", "") != ai_status:
                yn_cell.fill = changed_fill

        if not is_section:
            ws.row_dimensions[r].height = MIN_ROW_HEIGHT
        r += 1

    widths = {"A": 7.4, "B": 61.3, "C": 11.0, "D": 9.0, "E": 67.1}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_template_xlsx(items: list, header_labels: dict = None, note_text: str = "",
                        title: str = "||  JOB COMPLIANCE CHECKLIST (QC)  ||") -> bytes:
    """
    Build a read-only reference .xlsx for the admin Templates page download —
    shows Sno./Checklist Item/Reference/What to verify (no Yes/No or Remarks,
    since those are filled in later during an actual QC run, not here).
    """
    header_labels = header_labels or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    bold = Font(name="Calibri", size=11, bold=True)
    normal = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    section_fill  = PatternFill("solid", fgColor="E8590C")
    section_font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    ws.merge_cells("A1:E1")
    c = ws.cell(1, 1, title)
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A3:C3")
    ws.cell(3, 1, header_labels.get("customer_label") or "Customer Name  :").font = bold
    ws.merge_cells("A4:C4")
    ws.cell(4, 1, header_labels.get("address_label") or "Correct Address  :").font = bold
    ws.merge_cells("A5:C5")
    ws.cell(5, 1, "Checked By  :").font = bold
    ws.merge_cells("A6:C6")
    ws.cell(6, 1, "Date  :").font = bold
    ws.merge_cells("A7:E7")
    ws.cell(7, 1, header_labels.get("job_label") or "Job Details  :").font = bold

    headers = ["Sno.", "Checklist Item", "Reference", "What to verify as per Agreement"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(8, ci, h)
        cell.font = bold
        cell.alignment = center
        cell.border = BORDER
    ws.row_dimensions[8].height = 20

    r = 9
    MIN_ROW_HEIGHT = 30

    for it in items:
        if not it.get("active", 1):
            continue
        is_section = it.get("is_section")

        sno_cell    = ws.cell(r, 1, it.get("sno", ""))
        text_cell   = ws.cell(r, 2, it.get("text", ""))
        ref_cell    = ws.cell(r, 3, it.get("reference", ""))
        prompt_cell = ws.cell(r, 4, it.get("prompt", ""))

        for cell in (sno_cell, text_cell, ref_cell, prompt_cell):
            cell.border = BORDER

        if is_section:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            text_cell.font   = section_font
            text_cell.fill   = section_fill
            sno_cell.fill    = section_fill
            ref_cell.fill    = section_fill
            prompt_cell.fill = section_fill
        else:
            sno_cell.font  = bold
            text_cell.font = normal

        sno_cell.alignment    = center
        text_cell.alignment   = left
        ref_cell.alignment    = left
        prompt_cell.alignment = left

        if not is_section:
            ws.row_dimensions[r].height = MIN_ROW_HEIGHT
        r += 1

    widths = {"A": 7.4, "B": 61.3, "C": 30.0, "D": 67.1}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
