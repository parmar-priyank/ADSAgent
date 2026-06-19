"""
checklist_xlsx.py — parse a QC checklist .xlsx into structured items, and
regenerate a styled .xlsx from items (used for download after editing).

The regenerated file is freshly styled to closely match the original layout
(Calibri, bold title, header row, Sno/CheckList/Yes-No/Remarks columns, merged
remark cells, column widths). Because items can be added / removed / reordered,
the structured items are the source of truth and the file is rebuilt each time.
"""
import io
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

ROMAN = re.compile(r"^(?:i{1,3}|iv|v|vi{1,3}|ix|x)$", re.IGNORECASE)
HEADER_ROW_TEXT = ("sno", "checklist")
SECTION_HINT = re.compile(r"DETAILS|CEC|APPROVED|SYSTEM|CUSTOMER", re.IGNORECASE)

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_xlsx(file_bytes: bytes, sheet_name: str = None):
    """
    Read a QC checklist sheet and return (items, header_labels, note_text).
    items: list of {position, sno, parent_position, is_section, text, active}
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Pick the sheet with the most filled-in column-D (remarks) cells —
        # so a completed example sheet wins over a blank template.
        def score(s):
            return sum(1 for r in range(1, s.max_row + 1) if s.cell(r, 4).value)
        ws = max(wb.worksheets, key=score)

    note_text = ""
    items = []
    pos = 0
    last_numbered_pos = None

    # Find where the checklist table starts (the Sno./CheckList header row).
    start_row = None
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "").strip().lower()
        b = str(ws.cell(r, 2).value or "").strip().lower()
        if a.startswith("sno") and b.startswith("checklist"):
            start_row = r + 1
            break
    if start_row is None:
        start_row = 9  # sensible default for this template

    # Note text (column D of the title row, if any).
    note_cell = ws.cell(1, 4).value
    if note_cell and "note" not in str(note_cell).lower():
        note_text = str(note_cell)

    for r in range(start_row, ws.max_row + 1):
        sno = ws.cell(r, 1).value
        text = ws.cell(r, 2).value
        if text is None and sno is None:
            continue  # blank spacer row
        text = str(text).strip() if text is not None else ""
        sno_str = str(sno).strip() if sno is not None else ""
        if not text:
            continue

        # Reference text comes from column D (the Remarks column). Strip the
        # leading "Refer to " so the editor shows just the filename.
        remark = ws.cell(r, 4).value
        reference = ""
        if remark is not None:
            reference = re.sub(r"^\s*Refer to\s+", "", str(remark).strip(),
                               flags=re.IGNORECASE).strip()

        pos += 1
        is_section = bool(text.isupper() and SECTION_HINT.search(text)) or \
            (text.isupper() and len(text) > 4 and not sno_str.isdigit())

        is_sub = bool(ROMAN.match(sno_str)) or sno_str == ""
        parent = last_numbered_pos if (is_sub and not is_section) else None

        items.append({
            "position": pos,
            "sno": sno_str,
            "parent_position": parent,
            "is_section": is_section,
            "text": text,
            "reference": reference,
            "active": 1,
        })

        if sno_str.isdigit():
            last_numbered_pos = pos

    header_labels = {
        "customer_label": _label_at(ws, "customer"),
        "address_label": _label_at(ws, "address"),
        "job_label": _label_at(ws, "job details"),
    }
    return items, header_labels, note_text


def _label_at(ws, needle: str):
    for r in range(1, 12):
        v = str(ws.cell(r, 1).value or "")
        if needle.lower() in v.lower():
            # keep just the label part up to the colon
            return v.split(":")[0].strip() + "  :"
    return ""


def build_xlsx(items: list, header_labels: dict = None, note_text: str = "",
               title: str = "||  JOB COMPLIANCE CHECKLIST (QC)  ||",
               filled: dict = None) -> bytes:
    """
    Build a styled .xlsx from structured items.
    `filled` (optional): {position: {"status": "Y"/"N", "remark": "..."}}
    to populate the Yes/No and Remarks columns (used after a cross-check).
    """
    header_labels = header_labels or {}
    filled = filled or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    bold = Font(name="Calibri", size=11, bold=True)
    normal = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    section_fill = PatternFill("solid", fgColor="E8590C")
    section_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # Title row (merged A:E)
    ws.merge_cells("A1:E1")
    c = ws.cell(1, 1, title)
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 22

    # Header label rows
    ws.merge_cells("A3:C3")
    ws.cell(3, 1, header_labels.get("customer_label") or "Customer Name  :").font = bold
    ws.merge_cells("A4:C4")
    ws.cell(4, 1, header_labels.get("address_label") or "Correct Address  :").font = bold
    ws.merge_cells("A5:C5")
    ws.cell(5, 1, "Checked By  :").font = bold
    ws.merge_cells("A6:C6")
    ws.cell(6, 1, "Date  :").font = bold
    ws.merge_cells("A7:E7")
    ws.cell(7, 1, header_labels.get("job_label") or "Job Details  :").font = bold
    if note_text:
        ws.cell(1, 4)  # note kept simple; title already merged across

    # Column headers (row 8)
    headers = ["Sno.", "CheckList", "Yes/No", "Remarks"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(8, ci, h)
        cell.font = bold
        cell.alignment = center
        cell.border = BORDER
    ws.merge_cells("D8:E8")
    ws.cell(8, 4).border = BORDER
    ws.cell(8, 5).border = BORDER

    # Data rows
    r = 9
    for it in items:
        if not it.get("active", 1):
            continue
        pos = it.get("position")
        is_section = it.get("is_section")

        sno_cell = ws.cell(r, 1, it.get("sno", ""))
        text_cell = ws.cell(r, 2, it.get("text", ""))
        yn_cell = ws.cell(r, 3, "")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        remark_cell = ws.cell(r, 4, "")

        for cell in (sno_cell, text_cell, yn_cell, remark_cell, ws.cell(r, 5)):
            cell.border = BORDER

        if is_section:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
            text_cell.font = section_font
            text_cell.fill = section_fill
            sno_cell.fill = section_fill
            yn_cell.fill = section_fill
            remark_cell.fill = section_fill
            ws.cell(r, 5).fill = section_fill
        else:
            sno_cell.font = bold
            text_cell.font = normal

        sno_cell.alignment = center
        text_cell.alignment = left
        yn_cell.alignment = center
        remark_cell.alignment = left

        # populate Yes/No + remark if provided
        fv = filled.get(pos)
        if fv:
            yn_cell.value = fv.get("status", "")
            remark_cell.value = fv.get("remark", "")

        r += 1

    # Column widths (match original)
    widths = {"A": 7.4, "B": 61.3, "C": 11.0, "D": 9.0, "E": 67.1}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()